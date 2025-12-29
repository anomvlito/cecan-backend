import openpyxl
import os

# Load the Excel file
excel_path = r"2025_Cronograma Proy Desigualdades_ WP#1 CECAN_20032025 editado.xlsx"

if not os.path.exists(excel_path):
    print(f"File not found: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

print("="*80)
print("AUDITORÍA COMPLETA DEL EXCEL")
print("="*80)

# Analyze all rows
print("\nColumnas relevantes:")
print("A (col 0): Evento")
print("B (col 1): Evento instantaneo")
print("C (col 2): Producto al que pertenece")
print("D (col 3): Inicio")
print("E (col 4): Días")
print("F (col 5): Término")

print("\n" + "="*80)
print("TODOS LOS EVENTOS EN EL ARCHIVO")
print("="*80)

eventos_validos = []
eventos_invalidos = []
filas_vacias = []

for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), 2):
    col_a = row[0].value  # Evento
    col_b = row[1].value  # Evento instantaneo
    col_c = row[2].value  # Producto al que pertenece
    col_d = row[3].value  # Inicio
    col_e = row[4].value  # Días
    col_f = row[5].value  # Término
    
    # Check if row is completely empty
    if all(cell.value is None for cell in row[:6]):
        filas_vacias.append(i)
        continue
    
    # Check if it's a valid calendar event
    if col_a is not None and col_d is not None:
        eventos_validos.append({
            'fila': i,
            'evento': col_a,
            'instantaneo': col_b,
            'producto': col_c,
            'inicio': col_d,
            'dias': col_e,
            'termino': col_f
        })
    else:
        eventos_invalidos.append({
            'fila': i,
            'evento': col_a,
            'instantaneo': col_b,
            'producto': col_c,
            'inicio': col_d,
            'dias': col_e,
            'termino': col_f
        })

print(f"\nRESUMEN:")
print(f"  Total filas en archivo: {sheet.max_row - 1}")
print(f"  Eventos válidos (con nombre Y fecha inicio): {len(eventos_validos)}")
print(f"  Eventos inválidos/problemáticos: {len(eventos_invalidos)}")
print(f"  Filas vacías: {len(filas_vacias)}")

if eventos_validos:
    print(f"\nPRIMEROS 5 EVENTOS VÁLIDOS:")
    for ev in eventos_validos[:5]:
        print(f"  Fila {ev['fila']}: {ev['evento'][:50]}... | Inicio: {ev['inicio']} | Término: {ev['termino']}")

if eventos_invalidos:
    print(f"\nEVENTOS INVÁLIDOS/PROBLEMÁTICOS:")
    for ev in eventos_invalidos[:10]:
        print(f"  Fila {ev['fila']}: Evento={ev['evento']} | Inicio={ev['inicio']} | Término={ev['termino']}")

print("\n" + "="*80)
print("ANÁLISIS DE CONTENIDO NO-CALENDARIO")
print("="*80)

# Check for non-calendar items (headers, subtotals, etc.)
posibles_no_calendario = []

for ev in eventos_validos:
    nombre = str(ev['evento']).lower()
    # Check for patterns that might indicate non-calendar items
    if any(pattern in nombre for pattern in ['total', 'subtotal', 'resumen', 'suma', '===', '---']):
        posibles_no_calendario.append(ev)

if posibles_no_calendario:
    print(f"\nPOSIBLES ITEMS NO-CALENDARIO (subtotales, resúmenes, etc.):")
    for ev in posibles_no_calendario:
        print(f"  Fila {ev['fila']}: {ev['evento']}")
else:
    print("\nNo se detectaron patrones obvios de items no-calendario")


print("\n" + "="*80)
print("COLUMNA 'Evento instantaneo' - Análisis")
print("="*80)

instantaneos = [ev for ev in eventos_validos if ev['instantaneo'] is not None]
print(f"\nEventos con valor en 'Evento instantaneo': {len(instantaneos)}")
if instantaneos:
    print("Valores encontrados:")
    valores_unicos = set(ev['instantaneo'] for ev in instantaneos)
    for val in valores_unicos:
        count = sum(1 for ev in instantaneos if ev['instantaneo'] == val)
        print(f"  '{val}': {count} veces")
