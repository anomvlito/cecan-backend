import openpyxl
import os

# Load the Excel file
excel_path = r"2025_Cronograma Proy Desigualdades_ WP#1 CECAN_20032025 editado.xlsx"

if not os.path.exists(excel_path):
    print(f"File not found: {excel_path}")
    exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active

print(f"Sheet name: {sheet.title}")
print(f"Dimensions: {sheet.dimensions}")
print(f"Max row: {sheet.max_row}")
print(f"Max column: {sheet.max_column}")
print("\n" + "="*80)
print("HEADER DETECTION")
print("="*80)

# Check first 5 rows for headers
for row_idx in range(1, 6):
    row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]
    row_values = [c.value for c in row[:10]]  # First 10 columns
    print(f"Row {row_idx}: {row_values}")

print("\n" + "="*80)
print("ANALYZING COLUMNS")
print("="*80)

# Check what the parser is looking for
headers = {}
header_row_idx = 1

for row in sheet.iter_rows(min_row=1, max_row=5):
    row_values = [c.value for c in row]
    row_strings = [str(r).lower() if r else "" for r in row_values]
    
    print(f"Row {row[0].row} strings (first 10): {row_strings[:10]}")
    
    if "evento" in row_strings or "actividad" in row_strings:
        print(f"  -> Found header row at {row[0].row}")
        header_row_idx = row[0].row
        for cell in row:
            if not cell.value: continue
            val = str(cell.value).lower()
            if "evento" in val or "actividad" in val:
                headers["name"] = cell.column - 1
                print(f"  -> Name column: {cell.column} (0-indexed: {cell.column-1})")
            elif "inicio" in val or "start" in val:
                headers["start"] = cell.column - 1
                print(f"  -> Start column: {cell.column} (0-indexed: {cell.column-1})")
            elif "término" in val or "termino" in val or "fin" in val or "end" in val:
                headers["end"] = cell.column - 1
                print(f"  -> End column: {cell.column} (0-indexed: {cell.column-1})")
        break

print(f"\nHeaders detected: {headers}")
print(f"Header row index: {header_row_idx}")

print("\n" + "="*80)
print("SAMPLE DATA ROWS")
print("="*80)

# Show first 10 data rows
for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 10), 1):
    if "name" in headers and "start" in headers and "end" in headers:
        name_cell = row[headers["name"]]
        start_cell = row[headers["start"]]
        end_cell = row[headers["end"]]
        print(f"Row {header_row_idx + i}: Name={name_cell.value} | Start={start_cell.value} | End={end_cell.value}")
