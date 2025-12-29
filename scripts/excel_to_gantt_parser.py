
import os
import pandas as pd
import openpyxl
import hashlib
from datetime import timedelta
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelGanttParser:
    def __init__(self, file_path):
        self.file_path = file_path
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found at: {file_path}")

    def parse(self):
        """
        Parses the Excel file and returns a list of tasks in Frappe Gantt format.
        """
        try:
            # Load with openpyxl for colors
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = wb.active
            
            # Identify columns
            headers = {}
            header_row_idx = 1
            
            # Header detection with priority for exact matches
            for row in sheet.iter_rows(min_row=1, max_row=5):
                row_values = [c.value for c in row]
                row_strings = [str(r).lower() if r else "" for r in row_values]
                
                if "evento" in row_strings or "actividad" in row_strings:
                    header_row_idx = row[0].row
                    
                    # First pass: Look for EXACT matches (prioritize these)
                    for cell in row:
                        if not cell.value: continue
                        val = str(cell.value).lower().strip()
                        
                        if val == "evento":
                            headers["name"] = cell.column - 1
                        elif val == "inicio":
                            headers["start"] = cell.column - 1
                        elif val == "término":
                            headers["end"] = cell.column - 1
                        elif val == "días":
                            headers["duration"] = cell.column - 1
                        elif "producto" in val:
                            headers["product"] = cell.column - 1
                    
                    # Second pass: Look for partial matches only if not found
                    if "name" not in headers:
                        for cell in row:
                            if not cell.value: continue
                            val = str(cell.value).lower()
                            if "evento" in val or "actividad" in val:
                                headers["name"] = cell.column - 1
                                break
                    
                    if "start" not in headers:
                        for cell in row:
                            if not cell.value: continue
                            val = str(cell.value).lower()
                            if "inicio" in val or "start" in val:
                                headers["start"] = cell.column - 1
                                break
                    
                    if "end" not in headers:
                        for cell in row:
                            if not cell.value: continue
                            val = str(cell.value).lower()
                            if "término" in val or "termino" in val or "fin" in val or "end" in val:
                                headers["end"] = cell.column - 1
                                break
                    
                    if "duration" not in headers:
                        for cell in row:
                            if not cell.value: continue
                            val = str(cell.value).lower()
                            if "días" in val or "dias" in val or "duration" in val:
                                headers["duration"] = cell.column - 1
                                break
                    
                    if "product" not in headers:
                        for cell in row:
                            if not cell.value: continue
                            val = str(cell.value).lower()
                            if "producto" in val:
                                headers["product"] = cell.column - 1
                                break
                    
                    break
            
            if "name" not in headers or "start" not in headers or "end" not in headers:
                # Fallback to standard columns A, B, C if detection fails
                logger.warning("Could not detect headers automatically. Assuming A=Name, B=Start, C=End.")
                headers = {"name": 0, "start": 1, "end": 2}

            tasks = []
            
            # Iterate rows starting after header
            for row in sheet.iter_rows(min_row=header_row_idx + 1):
                name_cell = row[headers["name"]]
                start_cell = row[headers["start"]]
                end_cell = row[headers["end"]]
                
                if not name_cell.value:
                    continue
                
                # Extract product ID (default to 0 if not found)
                product_id = 0
                if "product" in headers:
                    product_cell = row[headers["product"]]
                    if product_cell.value is not None:
                        try:
                            product_id = int(product_cell.value)
                        except (ValueError, TypeError):
                            product_id = 0
                
                # Extract duration (days)
                duration = None
                if "duration" in headers:
                    duration_cell = row[headers["duration"]]
                    if duration_cell.value is not None:
                        try:
                            duration = int(duration_cell.value)
                        except (ValueError, TypeError):
                            duration = None
                
                # Check for NaT or basic validity using logic later, just capture raw now
                task = {
                    "name": name_cell.value,
                    "start": start_cell.value,
                    "end": end_cell.value,
                    "duration": duration,
                    "product_id": product_id,
                    "raw_id": f"{name_cell.row}" # Temporary ID
                }
                tasks.append(task)
            
            # Process with Pandas for easier date handling
            df = pd.DataFrame(tasks)
            
            # Normalize dates
            df['start'] = pd.to_datetime(df['start'], errors='coerce')
            df['end'] = pd.to_datetime(df['end'], errors='coerce')
            
            # Log events without start dates before filtering
            invalid_events = df[df['start'].isna()]
            if not invalid_events.empty:
                logger.warning(f"Found {len(invalid_events)} events without start dates (will be excluded from Gantt):")
                for _, event in invalid_events.iterrows():
                    logger.warning(f"  - {event['name']}")
            
            # Filter invalid dates
            df = df.dropna(subset=['start'])
            
            # Calculate end date from start + duration if end is missing
            for idx, row in df.iterrows():
                if pd.isna(row['end']) and row['duration'] is not None:
                    df.at[idx, 'end'] = row['start'] + timedelta(days=row['duration'])
            
            # If end is still missing, assume same day as start
            df['end'] = df['end'].fillna(df['start'])
            
            # Sort by start date
            df = df.sort_values(by='start')
            
            # Generate stable IDs (using enumerate to get row index for uniqueness)
            if not df.empty:
                df = df.reset_index(drop=True)  # Reset index to ensure sequential numbering
                df['id'] = [self._generate_id(row['name'], row['start'], idx) for idx, row in df.iterrows()]
            else:
                df['id'] = []
            
            # Product ID to CSS class mapping
            PRODUCT_CLASS_MAP = {
                0: "gantt-product-0",  # Azul (Base/Gestión)
                1: "gantt-product-1",  # Verde (Producto 1)
                2: "gantt-product-2",  # Amarillo (Producto 2)
                3: "gantt-product-3",  # Naranja (Producto 3)
                4: "gantt-product-4",  # Rojo (Producto 4)
                5: "gantt-product-5",  # morado (Producto 5)
            }
            
            # Build final task list
            final_tasks = []
            records = df.to_dict('records')
            
            for task in records:
                # Map product ID to CSS class
                product_id = task.get('product_id', 0)
                custom_class = PRODUCT_CLASS_MAP.get(product_id, "gantt-product-0")
                
                gantt_task = {
                    "id": task['id'],
                    "name": task['name'],
                    "start": task['start'].strftime('%Y-%m-%d'),
                    "end": task['end'].strftime('%Y-%m-%d'),
                    "dependencies": "",  # Empty by default - user will create manually
                    "custom_class": custom_class
                }
                
                final_tasks.append(gantt_task)
            
            return final_tasks
            
        except Exception as e:
            logger.error(f"Error parsing Excel: {e}")
            raise e

    def _generate_id(self, name, start_date, row_number):
        # Create a deterministic hash based on name, start date, and row number for uniqueness
        s = f"{name}_{start_date.isoformat()}_{row_number}".encode('utf-8')
        return hashlib.md5(s).hexdigest()[:10]

if __name__ == "__main__":
    # Test execution
    import sys
    # Adjust path to find the file if running from script dir
    # Assuming script is in backend/scripts/ and excel is in root
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    excel_path = os.path.join(base_dir, "2025_Cronograma Proy Desigualdades_ WP#1 CECAN_20032025 editado.xlsx")
    
    print(f"Looking for file at: {excel_path}")
    if os.path.exists(excel_path):
        parser = ExcelGanttParser(excel_path)
        tasks = parser.parse()
        import json
        print(json.dumps(tasks, indent=2))
    else:
        print("File not found for testing.")
